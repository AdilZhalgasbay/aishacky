import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook("для хакатона расписание.xlsx", data_only=True)
ws = wb["сабақ кестесі"]

print(f"Размер листа: {ws.max_row} строк x {ws.max_column} колонок\n")

# Печатаем первые 30 строк
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True)):
    cells = []
    for c in row:
        v = str(c).replace('\xa0','').strip() if c is not None else ''
        cells.append(v[:25] if v else '.')
    print(f"R{i+1:02d}: {cells}")
