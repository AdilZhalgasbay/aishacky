import openpyxl
import os
import sys

# Force UTF-8 output on Windows
sys.stdout.reconfigure(encoding='utf-8')

files = [
    "для хакатона расписание.xlsx",
    "нагрузка учителей для хакатона 2025-2026.xlsx",
]

for fname in files:
    if not os.path.exists(fname):
        print(f"NOT FOUND: {fname}")
        continue
    print(f"\n{'='*60}")
    print(f"FILE: {fname}")
    print('='*60)
    wb = openpyxl.load_workbook(fname, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  SHEET: '{sheet_name}'  ({ws.max_row} rows x {ws.max_column} cols)")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 15:
                print(f"  ... ({ws.max_row - 15} more rows)")
                break
            if any(c is not None for c in row):
                # Truncate long cells
                cells = [str(c)[:40] if c is not None else None for c in row]
                print(f"  Row {i+1}: {cells}")
