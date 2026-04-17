"""
import_xlsx.py — Импорт реальных данных школы Aqbobek в Supabase
Таблицы: employees, classes, schedules
"""
import sys, re, requests, json
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from supabase import create_client

SUPABASE_URL = "https://tutzawhhpklqodjagtha.supabase.co"
SERVICE_KEY  = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InR1dHphd2hocGtscW9kamFndGhhIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjQzNDMyNSwiZXhwIjoyMDkyMDEwMzI1fQ.jcHpPsu6s1h-lzycNYZ8oAheCE2kLVBer1KSbkN7J0o"
MGMT_TOKEN   = "sbp_9f5df5a5da0af2f9551ae2bb078f10bed0d2154f"
PROJECT      = "tutzawhhpklqodjagtha"

sb = create_client(SUPABASE_URL, SERVICE_KEY)

def clean(v):
    if v is None: return None
    s = str(v).replace('\xa0', '').strip()
    return s if s else None

def intval(v):
    s = clean(v)
    if not s: return None
    try: return int(float(s))
    except: return None

# ────────────────────────────────────────────────────────────────────────────
# 1. УЧИТЕЛЯ из "нагрузка учителей для хакатона 2025-2026.xlsx"
# ────────────────────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("1. ИМПОРТ УЧИТЕЛЕЙ")
print("="*60)

wb = openpyxl.load_workbook("нагрузка учителей для хакатона 2025-2026.xlsx", data_only=True)
ws = wb["Жүктеме 2025-2026"]

teachers = {}   # name -> {subjects: set}

current_name = None
for row in ws.iter_rows(values_only=True):
    num  = clean(row[0])
    name = clean(row[1])
    subj = clean(row[3])

    if num and re.match(r'^\d+$', num) and name and len(name) > 3 and name != '0':
        current_name = name
        if current_name not in teachers:
            teachers[current_name] = {"subjects": set()}

    if current_name and subj and subj not in {'оқушылар саны', 'Диплом бойынша мамандығы'}:
        skip_subjects = {'ҰБТ', 'мат.сау', 'қосымша', 'ДЧ', 'спец.курс'}
        if subj not in skip_subjects:
            teachers[current_name]["subjects"].add(subj)

print(f"Найдено учителей в файле: {len(teachers)}")

# Очищаем старых моковых сотрудников и вставляем реальных
# Сначала получаем существующих
existing_res = sb.table("employees").select("id, name").execute()
existing_map = {e["name"]: e["id"] for e in existing_res.data}
print(f"Уже в БД: {len(existing_map)} сотрудников")

inserted = 0
updated  = 0
teacher_id_map = {}  # name -> uuid

for name, info in teachers.items():
    subj_str = ", ".join(sorted(info["subjects"]))[:200] if info["subjects"] else None

    if name in existing_map:
        emp_id = existing_map[name]
        sb.table("employees").update({
            "subject": subj_str,
            "role": "teacher",
            "is_available": True,
        }).eq("id", emp_id).execute()
        teacher_id_map[name] = emp_id
        updated += 1
    else:
        res = sb.table("employees").insert({
            "name": name,
            "role": "teacher",
            "subject": subj_str,
            "is_available": True,
        }).execute()
        if res.data:
            teacher_id_map[name] = res.data[0]["id"]
            inserted += 1

print(f"Вставлено: {inserted}, Обновлено: {updated}")
print(f"Учителей в маппинге: {len(teacher_id_map)}")

# ────────────────────────────────────────────────────────────────────────────
# 2. КАБИНЕТЫ / КЛАССЫ из листа "Кабинеттер тізімі"
# ────────────────────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("2. ИМПОРТ КЛАССОВ")
print("="*60)

ws_rooms = wb["Кабинеттер тізімі"]

# Сначала очищаем
sb.table("classes").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()

classes_inserted = 0
class_id_map = {}  # class_name -> uuid

for row in ws_rooms.iter_rows(min_row=2, values_only=True):
    room_num    = clean(row[1])
    floor       = clean(row[2])
    capacity    = clean(row[3])
    class_name  = clean(row[4])
    teacher_name = clean(row[5])

    if not room_num:
        continue

    # Пропускаем кабинеты без класса (лаборатории, спортзал и т.д.)
    if not class_name:
        continue

    # Определяем grade из имени класса (7A -> 7)
    grade = 0  # default
    m = re.match(r'^(\d+)', class_name)
    if m:
        grade = int(m.group(1))

    teacher_id = None
    if teacher_name:
        # Ищем учителя по части имени (может быть несколько через \n)
        for part in re.split(r'[\n,]', teacher_name):
            part = part.strip()
            for tname, tid in teacher_id_map.items():
                if part and (part in tname or tname.split()[0] in part):
                    teacher_id = tid
                    break
            if teacher_id:
                break

    res = sb.table("classes").insert({
        "name": class_name,
        "grade": grade,
        "room_number": room_num,
        "student_count": intval(capacity),
        "homeroom_teacher_id": teacher_id,
    }).execute()

    if res.data:
        cname = class_name or f"Каб.{room_num}"
        class_id_map[cname] = res.data[0]["id"]
        classes_inserted += 1

print(f"Классов/кабинетов вставлено: {classes_inserted}")

# ────────────────────────────────────────────────────────────────────────────
# 3. РАСПИСАНИЕ из "для хакатона расписание.xlsx"
# ────────────────────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("3. ИМПОРТ РАСПИСАНИЯ")
print("="*60)

wb2 = openpyxl.load_workbook("для хакатона расписание.xlsx", data_only=True)

day_map = {
    'дүйсенбі': 1, 'понедельник': 1,
    'сейсенбі': 2, 'вторник': 2,
    'сәрсенбі': 3, 'среда': 3,
    'бейсенбі': 4, 'четверг': 4,
    'жұма': 5, 'пятница': 5,
}

# Очищаем старое расписание
sb.table("schedules").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()

sched_rows = []

for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]

    # Пробуем распарсить лист как расписание:
    # Ищем строки где есть: имя учителя / предмет / класс / день / урок
    rows_data = list(ws2.iter_rows(values_only=True))
    if len(rows_data) < 3:
        continue

    # Попробуем найти заголовки в первых 5 строках
    headers = []
    header_row_idx = 0
    for i, row in enumerate(rows_data[:8]):
        non_null = [clean(c) for c in row if clean(c)]
        if len(non_null) >= 4:
            headers = [clean(c) for c in row]
            header_row_idx = i
            break

    if not headers:
        print(f"  Пропускаем '{sheet_name}' — нет заголовков")
        continue

    print(f"  Sheet '{sheet_name}': заголовки = {headers[:8]}")

    # Определяем индексы колонок
    def find_col(keywords):
        for kw in keywords:
            for j, h in enumerate(headers):
                if h and kw.lower() in str(h).lower():
                    return j
        return None

    col_teacher = find_col(['мұғалім', 'учитель', 'teacher', 'ФИО', 'аты'])
    col_subject = find_col(['пән', 'предмет', 'subject'])
    col_class   = find_col(['сынып', 'класс', 'class'])
    col_day     = find_col(['күн', 'день', 'day'])
    col_period  = find_col(['сабақ', 'урок', 'period', '№'])
    col_room    = find_col(['каб', 'room', 'кабинет'])

    print(f"    Колонки: учитель={col_teacher}, предмет={col_subject}, класс={col_class}, день={col_day}, урок={col_period}")

    if col_teacher is None and col_subject is None:
        print(f"    Пропускаем — нет нужных колонок")
        continue

    for row in rows_data[header_row_idx+1:]:
        teacher_name = clean(row[col_teacher]) if col_teacher is not None else None
        subject      = clean(row[col_subject]) if col_subject is not None else None
        class_name   = clean(row[col_class])   if col_class  is not None else None
        day_raw      = clean(row[col_day])      if col_day    is not None else None
        period_raw   = clean(row[col_period])   if col_period is not None else None
        room         = clean(row[col_room])     if col_room   is not None else None

        if not subject or not teacher_name:
            continue

        day_num = day_map.get(day_raw.lower() if day_raw else '', None)
        period_num = intval(period_raw)

        # Найти teacher_id
        emp_id = None
        if teacher_name:
            for tname, tid in teacher_id_map.items():
                if teacher_name in tname or tname in teacher_name or \
                   teacher_name.split()[0] in tname:
                    emp_id = tid
                    break

        # Найти class_id
        cls_id = class_id_map.get(class_name) if class_name else None

        sched_rows.append({
            "employee_id": emp_id,
            "class_id":    cls_id,
            "subject":     subject[:100],
            "day_of_week": day_num or 1,
            "period":      period_num or 1,
            "room":        room,
        })

# Batch insert по 100
if sched_rows:
    batch_size = 100
    for i in range(0, len(sched_rows), batch_size):
        batch = sched_rows[i:i+batch_size]
        sb.table("schedules").insert(batch).execute()
    print(f"Расписания вставлено: {len(sched_rows)} записей")
else:
    print("Расписание не распарсилось автоматически (сложная структура Excel)")
    print("Вставим демо-данные на основе учителей...")

    # Fallback: создаём демо расписание на основе нагрузки учителей
    demo_sched = []
    day = 1
    period = 1
    for name, info in list(teachers.items())[:15]:
        emp_id = teacher_id_map.get(name)
        if not emp_id:
            continue
        for subj in list(info["subjects"])[:2]:
            # Найти класс для этого учителя
            cls_id = None
            for cname, cid in class_id_map.items():
                if cname and len(cname) <= 4:  # 7A, 8B etc
                    cls_id = cid
                    break

            demo_sched.append({
                "employee_id": emp_id,
                "class_id":    cls_id,
                "subject":     subj[:100],
                "day_of_week": day,
                "period":      period,
                "room":        None,
            })
            period += 1
            if period > 7:
                period = 1
                day = (day % 5) + 1

    if demo_sched:
        for i in range(0, len(demo_sched), 100):
            sb.table("schedules").insert(demo_sched[i:i+100]).execute()
        print(f"Demo расписания вставлено: {len(demo_sched)} записей")

print("\n" + "="*60)
print("ГОТОВО! Данные загружены в Supabase.")
print("="*60)

# Итоговая статистика
r1 = sb.table("employees").select("id", count="exact").execute()
r2 = sb.table("classes").select("id", count="exact").execute()
r3 = sb.table("schedules").select("id", count="exact").execute()
print(f"\nСотрудников в БД:  {r1.count}")
print(f"Классов в БД:      {r2.count}")
print(f"Расписаний в БД:   {r3.count}")
